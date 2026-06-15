"""
MRS Payroll Anomaly Detector
Master Roofing Solutions | Payroll Validation Script
"""

from pathlib import Path
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList

# ─── PATHS ───────────────────────────────────────────────────────────────────

INPUT_PATH  = Path("/home/claude/mrs_payroll_raw.csv")
OUTPUT_PATH = Path("/home/claude/MRS_Payroll_Validation_Report.xlsx")

# ─── CONSTANTS ───────────────────────────────────────────────────────────────

VALID_PAY_TYPES = {"hourly", "piece_rate", "salary"}
MIN_WAGE        = {"CA": 16.50, "TX": 7.25, "FL": 13.00, "NY": 16.00, "WA": 16.28}
FLSA_OT_THRESHOLD   = 40.0
CA_DAILY_OT_THRESHOLD = 8.0
CA_DAILY_DT_THRESHOLD = 12.0
CA_MEAL_BREAK_TRIGGER = 5.0

# ─── BRAND COLORS (navy/teal MRS palette) ────────────────────────────────────

NAVY  = "1B2A4A"
TEAL  = "2A7F7F"
WHITE = "FFFFFF"
LIGHT_GRAY = "F2F4F7"
RED_ERR    = "C0392B"
AMBER_WARN = "E67E22"
GREEN_OK   = "27AE60"
LIGHT_RED  = "FADBD8"
LIGHT_AMBER= "FDEBD0"
LIGHT_GREEN= "D5F5E3"

# ─── HELPERS ─────────────────────────────────────────────────────────────────

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def header_font(color=WHITE, bold=True, size=10):
    return Font(name="Arial", bold=bold, color=color, size=size)

def cell_font(bold=False, size=10, color="000000"):
    return Font(name="Arial", bold=bold, size=size, color=color)

def fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def set_col_widths(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def style_header_row(ws, row, bg=NAVY, fg=WHITE, height=22):
    ws.row_dimensions[row].height = height
    for cell in ws[row]:
        if cell.value is not None:
            cell.font      = header_font(color=fg)
            cell.fill      = fill(bg)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = thin_border()

def style_data_row(ws, row_num, n_cols, bg=WHITE):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill      = fill(bg)
        cell.font      = cell_font()
        cell.alignment = Alignment(horizontal="left", vertical="center",
                                   wrap_text=False)
        cell.border    = thin_border()

# ─── LOAD DATA ───────────────────────────────────────────────────────────────

def load_data(path: Path) -> pd.DataFrame:
    df = pd.read_csv(path, parse_dates=["week_ending"])
    df["week_ending"] = pd.to_datetime(df["week_ending"]).dt.date
    return df

# ─── VALIDATION ENGINE ───────────────────────────────────────────────────────

def run_validations(df: pd.DataFrame) -> pd.DataFrame:
    flags = []

    for idx, row in df.iterrows():
        eid   = row["employee_id"]
        name  = row["employee_name"]
        state = row["state"]
        ptype = row["pay_type"]
        week  = row["week_ending"]
        reg   = row["regular_hours"]
        ot    = row["ot_hours"]
        dt    = row["double_time_hours"]
        shift = row["shift_hours"]
        hr    = row["hourly_rate"]
        pr    = row["piece_rate_usd"]
        pcs   = row["pieces_completed"]
        meal  = row["meal_break_taken"]
        mw    = MIN_WAGE.get(state, 7.25)

        def flag(rule_id, severity, description, impact):
            flags.append({
                "employee_id":   eid,
                "employee_name": name,
                "state":         state,
                "pay_type":      ptype,
                "week_ending":   week,
                "rule_id":       rule_id,
                "severity":      severity,
                "description":   description,
                "potential_impact": impact,
            })

        # R02 — duplicate check handled separately (needs group context)

        # R03 — missing/zero rates
        if ptype in ("hourly", "salary"):
            if pd.isna(hr) or hr == 0:
                flag("R03", "CRITICAL",
                     f"hourly_rate is {'null' if pd.isna(hr) else '0'} — employee has no pay rate configured",
                     "Employee would receive $0 gross pay")

        if ptype == "piece_rate":
            if pd.isna(pr) or pr == 0:
                flag("R03", "CRITICAL",
                     f"piece_rate_usd is {'null' if pd.isna(pr) else '0'} — piece rate not configured",
                     "Gross pay uncalculable; potential underpayment")

        # R04 — invalid pay_type
        if ptype not in VALID_PAY_TYPES:
            flag("R04", "HIGH",
                 f"pay_type = '{ptype}' is not a valid value (expected: hourly, piece_rate, salary)",
                 "Payroll system may reject record or miscalculate taxes")

        # R05 — FLSA OT: total hours > 40 but ot_hours = 0
        total_hours = reg + ot
        if total_hours > FLSA_OT_THRESHOLD and ot == 0:
            ot_owed = round(total_hours - FLSA_OT_THRESHOLD, 2)
            flag("R05", "CRITICAL",
                 f"Employee worked {total_hours}h but ot_hours = 0 — FLSA OT threshold breached",
                 f"~{ot_owed}h of OT unpaid; FLSA violation risk")

        # R06 — CA daily OT: shift > 8h, ot = 0
        if state == "CA" and shift > CA_DAILY_OT_THRESHOLD and ot == 0:
            flag("R06", "CRITICAL",
                 f"CA employee: shift_hours = {shift}h (>8h) but ot_hours = 0 — daily OT not captured",
                 "CA Labor Code §510 violation; IWC Wage Order exposure")

        # R07 — CA double time: shift > 12h, dt = 0
        if state == "CA" and shift > CA_DAILY_DT_THRESHOLD and dt == 0:
            flag("R07", "HIGH",
                 f"CA employee: shift_hours = {shift}h (>12h) but double_time_hours = 0",
                 "CA Labor Code §510 double-time not applied; underpayment")

        # R08 — piece-rate minimum wage floor
        if ptype == "piece_rate" and not pd.isna(pr) and pr > 0 and pcs and shift > 0:
            effective_rate = (pcs * pr) / shift
            if effective_rate < mw:
                shortfall = round((mw - effective_rate) * shift, 2)
                flag("R08", "CRITICAL",
                     f"Piece-rate effective hourly = ${effective_rate:.2f}/hr < {state} min wage ${mw}/hr",
                     f"Minimum wage breach; ~${shortfall} supplement owed this shift")

        # R10 — CA meal break premium
        if state == "CA" and shift > CA_MEAL_BREAK_TRIGGER and meal == False:
            flag("R10", "HIGH",
                 f"CA employee: shift = {shift}h, meal break NOT taken — 1hr premium owed",
                 f"~${mw:.2f} meal break premium per CA Labor Code §512")

        # R11 — CA second meal break
        if state == "CA" and shift > 10 and meal == False:
            flag("R11", "MEDIUM",
                 f"CA employee: shift = {shift}h (>10h), meal break not taken — 2nd premium may apply",
                 f"Additional ${mw:.2f} if 2nd meal break also missed")

    # R02 — duplicate detection (group-level)
    dupes = df[df.duplicated(subset=["employee_id", "week_ending"], keep=False)]
    for _, row in dupes.iterrows():
        flags.append({
            "employee_id":      row["employee_id"],
            "employee_name":    row["employee_name"],
            "state":            row["state"],
            "pay_type":         row["pay_type"],
            "week_ending":      row["week_ending"],
            "rule_id":          "R02",
            "severity":         "CRITICAL",
            "description":      f"Duplicate record: same employee_id + week_ending appears more than once",
            "potential_impact": "Risk of double payment; must be resolved before processing",
        })

    return pd.DataFrame(flags).drop_duplicates()

# ─── SHEET 1: SUMMARY ────────────────────────────────────────────────────────

def build_summary(wb: Workbook, df: pd.DataFrame, flags: pd.DataFrame):
    ws = wb.active
    ws.title = "1 · Summary"
    ws.sheet_view.showGridLines = False

    # Title block
    ws.merge_cells("A1:H1")
    ws["A1"] = "MASTER ROOFING SOLUTIONS — PAYROLL VALIDATION REPORT"
    ws["A1"].font      = Font(name="Arial", bold=True, size=14, color=WHITE)
    ws["A1"].fill      = fill(NAVY)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:H2")
    ws["A2"] = f"Period: {min(df['week_ending'])} – {max(df['week_ending'])}   |   Generated: {pd.Timestamp.today().strftime('%Y-%m-%d')}"
    ws["A2"].font      = Font(name="Arial", size=10, color=WHITE)
    ws["A2"].fill      = fill(TEAL)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # KPI Cards (row 4-7)
    total_records  = len(df)
    total_employees= df["employee_id"].nunique()
    total_flags    = len(flags)
    clean_records  = total_records - flags["employee_id"].nunique()
    critical_count = len(flags[flags["severity"] == "CRITICAL"])
    high_count     = len(flags[flags["severity"] == "HIGH"])
    medium_count   = len(flags[flags["severity"] == "MEDIUM"])

    kpis = [
        ("Total Records",    total_records,   NAVY,     "A"),
        ("Employees",        total_employees, TEAL,     "C"),
        ("Total Flags",      total_flags,     RED_ERR,  "E"),
        ("Clean Records",    clean_records,   GREEN_OK, "G"),
    ]

    for label, value, color, col in kpis:
        # label row
        lc = ws[f"{col}4"]
        lc.value     = label
        lc.font      = Font(name="Arial", bold=True, size=9, color=WHITE)
        lc.fill      = fill(color)
        lc.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(f"{col}4:{chr(ord(col)+1)}4")
        # value row
        vc = ws[f"{col}5"]
        vc.value     = value
        vc.font      = Font(name="Arial", bold=True, size=22, color=color)
        vc.fill      = fill(WHITE)
        vc.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(f"{col}5:{chr(ord(col)+1)}6")
        ws.row_dimensions[4].height = 18
        ws.row_dimensions[5].height = 36
        ws.row_dimensions[6].height = 36

    # Severity breakdown (row 8-9)
    ws.row_dimensions[8].height = 6  # spacer

    sev_data = [
        ("CRITICAL", critical_count, RED_ERR,    LIGHT_RED,   "B"),
        ("HIGH",     high_count,     AMBER_WARN,  LIGHT_AMBER, "D"),
        ("MEDIUM",   medium_count,   TEAL,        LIGHT_GREEN, "F"),
    ]
    ws["A9"] = "Severity Breakdown"
    ws["A9"].font = Font(name="Arial", bold=True, size=10, color=NAVY)

    for sev, count, badge_color, bg_color, col in sev_data:
        c1 = ws[f"{col}9"]
        c1.value     = f"  {sev}  "
        c1.font      = Font(name="Arial", bold=True, size=9, color=WHITE)
        c1.fill      = fill(badge_color)
        c1.alignment = Alignment(horizontal="center")
        ws.merge_cells(f"{col}9:{chr(ord(col)+1)}9")

        c2 = ws[f"{col}10"]
        c2.value     = count
        c2.font      = Font(name="Arial", bold=True, size=16, color=badge_color)
        c2.fill      = fill(bg_color)
        c2.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(f"{col}10:{chr(ord(col)+1)}10")
        ws.row_dimensions[9].height  = 18
        ws.row_dimensions[10].height = 28

    # Flags by Rule table (row 12+)
    ws.row_dimensions[11].height = 8  # spacer
    ws["A12"] = "Flags by Rule"
    ws["A12"].font = Font(name="Arial", bold=True, size=10, color=NAVY)

    rule_summary = (
        flags.groupby(["rule_id", "severity"])
        .size()
        .reset_index(name="count")
        .sort_values("rule_id")
    )

    rule_descriptions = {
        "R02": "Duplicate record (same employee + week)",
        "R03": "Missing or zero pay rate",
        "R04": "Invalid pay_type value",
        "R05": "FLSA weekly OT not captured",
        "R06": "CA daily OT not captured (>8h shift)",
        "R07": "CA double time not applied (>12h shift)",
        "R08": "Piece-rate effective hourly < state minimum wage",
        "R10": "CA meal break premium owed (1st break)",
        "R11": "CA 2nd meal break premium may apply",
    }

    headers = ["Rule", "Severity", "Description", "Flag Count"]
    ws.append([])
    hdr_row = 13
    for col_i, h in enumerate(headers, 1):
        c = ws.cell(row=hdr_row, column=col_i, value=h)
        c.font      = header_font()
        c.fill      = fill(NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = thin_border()
    ws.row_dimensions[hdr_row].height = 20

    sev_colors = {"CRITICAL": (RED_ERR, LIGHT_RED),
                  "HIGH":     (AMBER_WARN, LIGHT_AMBER),
                  "MEDIUM":   (TEAL, LIGHT_GREEN)}

    for i, (_, r) in enumerate(rule_summary.iterrows()):
        row_n = hdr_row + 1 + i
        badge_c, bg_c = sev_colors.get(r["severity"], (NAVY, LIGHT_GRAY))
        row_bg = LIGHT_GRAY if i % 2 == 0 else WHITE
        vals = [r["rule_id"], r["severity"],
                rule_descriptions.get(r["rule_id"], ""), r["count"]]
        for col_i, val in enumerate(vals, 1):
            c = ws.cell(row=row_n, column=col_i, value=val)
            c.font   = cell_font(bold=(col_i == 2), color=(badge_c if col_i == 2 else "000000"))
            c.fill   = fill(bg_c if col_i == 2 else row_bg)
            c.border = thin_border()
            c.alignment = Alignment(horizontal=("center" if col_i in (1,2,4) else "left"),
                                    vertical="center")
        ws.row_dimensions[row_n].height = 18

    set_col_widths(ws, {"A":10,"B":12,"C":52,"D":12,"E":12,"F":12,"G":12,"H":12})

# ─── SHEET 2: ANOMALIES ──────────────────────────────────────────────────────

def build_anomalies(wb: Workbook, flags: pd.DataFrame):
    ws = wb.create_sheet("2 · Anomalies")
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:I1")
    ws["A1"] = "ANOMALY LOG — All Validation Flags"
    ws["A1"].font      = Font(name="Arial", bold=True, size=12, color=WHITE)
    ws["A1"].fill      = fill(NAVY)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    headers = ["Employee ID", "Name", "State", "Pay Type",
               "Week Ending", "Rule", "Severity", "Description", "Potential Impact"]
    for col_i, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col_i, value=h)
        c.font      = header_font()
        c.fill      = fill(TEAL)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = thin_border()
    ws.row_dimensions[2].height = 20

    sev_bg = {"CRITICAL": LIGHT_RED, "HIGH": LIGHT_AMBER, "MEDIUM": LIGHT_GREEN}
    sev_fg = {"CRITICAL": RED_ERR,   "HIGH": AMBER_WARN,  "MEDIUM": TEAL}

    flags_sorted = flags.sort_values(["severity", "rule_id", "employee_id"])

    for i, (_, r) in enumerate(flags_sorted.iterrows()):
        row_n = i + 3
        bg    = sev_bg.get(r["severity"], WHITE)
        fg    = sev_fg.get(r["severity"], NAVY)
        row_bg= LIGHT_GRAY if i % 2 == 0 else WHITE

        vals = [r["employee_id"], r["employee_name"], r["state"], r["pay_type"],
                str(r["week_ending"]), r["rule_id"], r["severity"],
                r["description"], r["potential_impact"]]

        for col_i, val in enumerate(vals, 1):
            c = ws.cell(row=row_n, column=col_i, value=val)
            if col_i == 7:  # severity badge column
                c.font  = Font(name="Arial", bold=True, size=9, color=fg)
                c.fill  = fill(bg)
            else:
                c.font  = cell_font()
                c.fill  = fill(row_bg)
            c.border    = thin_border()
            c.alignment = Alignment(horizontal=("center" if col_i in (1,3,5,6,7) else "left"),
                                    vertical="center", wrap_text=(col_i in (8,9)))
        ws.row_dimensions[row_n].height = 30

    set_col_widths(ws, {
        "A":14, "B":22, "C":8, "D":12, "E":14,
        "F":8,  "G":12, "H":55, "I":45
    })

# ─── SHEET 3: DASHBOARD ──────────────────────────────────────────────────────

def build_dashboard(wb: Workbook, df: pd.DataFrame, flags: pd.DataFrame):
    ws = wb.create_sheet("3 · Dashboard")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:L1")
    ws["A1"] = "PAYROLL VALIDATION DASHBOARD — Master Roofing Solutions"
    ws["A1"].font      = Font(name="Arial", bold=True, size=13, color=WHITE)
    ws["A1"].fill      = fill(NAVY)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── Chart 1: Flags by Rule ──────────────────────────────────────────────
    rule_counts = flags.groupby("rule_id").size().reset_index(name="count").sort_values("rule_id")

    ws["A3"] = "Flags by Rule"
    ws["A3"].font = Font(name="Arial", bold=True, size=10, color=NAVY)
    ws["A4"] = "Rule"; ws["B4"] = "Count"
    for col_i in (1, 2):
        c = ws.cell(row=4, column=col_i)
        c.font  = header_font()
        c.fill  = fill(NAVY)
        c.alignment = Alignment(horizontal="center")

    for i, (_, r) in enumerate(rule_counts.iterrows()):
        ws.cell(row=5+i, column=1, value=r["rule_id"])
        ws.cell(row=5+i, column=2, value=int(r["count"]))

    chart1 = BarChart()
    chart1.type   = "col"
    chart1.title  = "Flags by Rule"
    chart1.y_axis.title = "Count"
    chart1.x_axis.title = "Rule ID"
    chart1.style  = 10
    chart1.width  = 14
    chart1.height = 10

    data1 = Reference(ws, min_col=2, min_row=4, max_row=4+len(rule_counts))
    cats1 = Reference(ws, min_col=1, min_row=5, max_row=4+len(rule_counts))
    chart1.add_data(data1, titles_from_data=True)
    chart1.set_categories(cats1)
    chart1.series[0].graphicalProperties.solidFill = NAVY
    ws.add_chart(chart1, "D3")

    # ── Chart 2: Flags by State ─────────────────────────────────────────────
    state_counts = flags.groupby("state").size().reset_index(name="count").sort_values("count", ascending=False)

    start_row = 5 + len(rule_counts) + 3
    ws.cell(row=start_row, column=1, value="State").font   = header_font()
    ws.cell(row=start_row, column=1).fill = fill(NAVY)
    ws.cell(row=start_row, column=2, value="Count").font  = header_font()
    ws.cell(row=start_row, column=2).fill = fill(NAVY)

    for i, (_, r) in enumerate(state_counts.iterrows()):
        ws.cell(row=start_row+1+i, column=1, value=r["state"])
        ws.cell(row=start_row+1+i, column=2, value=int(r["count"]))

    chart2 = BarChart()
    chart2.type   = "col"
    chart2.title  = "Flags by State"
    chart2.y_axis.title = "Count"
    chart2.x_axis.title = "State"
    chart2.style  = 10
    chart2.width  = 14
    chart2.height = 10

    data2 = Reference(ws, min_col=2, min_row=start_row, max_row=start_row+len(state_counts))
    cats2 = Reference(ws, min_col=1, min_row=start_row+1, max_row=start_row+len(state_counts))
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats2)
    chart2.series[0].graphicalProperties.solidFill = TEAL
    ws.add_chart(chart2, "D21")

    # ── Chart 3: Flags by Severity ──────────────────────────────────────────
    sev_counts = flags.groupby("severity").size().reset_index(name="count")
    sev_order  = ["CRITICAL", "HIGH", "MEDIUM"]
    sev_counts["severity"] = pd.Categorical(sev_counts["severity"], categories=sev_order, ordered=True)
    sev_counts = sev_counts.sort_values("severity")

    sev_row = start_row + len(state_counts) + 3
    ws.cell(row=sev_row, column=1, value="Severity").font  = header_font()
    ws.cell(row=sev_row, column=1).fill = fill(NAVY)
    ws.cell(row=sev_row, column=2, value="Count").font    = header_font()
    ws.cell(row=sev_row, column=2).fill = fill(NAVY)

    sev_fill_colors = {"CRITICAL": RED_ERR, "HIGH": AMBER_WARN, "MEDIUM": TEAL}
    for i, (_, r) in enumerate(sev_counts.iterrows()):
        ws.cell(row=sev_row+1+i, column=1, value=r["severity"])
        ws.cell(row=sev_row+1+i, column=2, value=int(r["count"]))

    chart3 = BarChart()
    chart3.type   = "bar"
    chart3.title  = "Flags by Severity"
    chart3.y_axis.title = "Severity"
    chart3.x_axis.title = "Count"
    chart3.style  = 10
    chart3.width  = 14
    chart3.height = 8

    data3 = Reference(ws, min_col=2, min_row=sev_row, max_row=sev_row+len(sev_counts))
    cats3 = Reference(ws, min_col=1, min_row=sev_row+1, max_row=sev_row+len(sev_counts))
    chart3.add_data(data3, titles_from_data=True)
    chart3.set_categories(cats3)
    chart3.series[0].graphicalProperties.solidFill = RED_ERR
    ws.add_chart(chart3, "D39")

    set_col_widths(ws, {"A":12, "B":10, "C":6})

# ─── MAIN ────────────────────────────────────────────────────────────────────

def main():
    df    = load_data(INPUT_PATH)
    flags = run_validations(df)

    wb = Workbook()
    build_summary(wb, df, flags)
    build_anomalies(wb, flags)
    build_dashboard(wb, df, flags)

    wb.save(OUTPUT_PATH)
    print(f"Report saved → {OUTPUT_PATH}")
    print(f"Total flags  : {len(flags)}")
    print(f"\nFlag breakdown:\n{flags['severity'].value_counts()}")
    print(f"\nFlags by rule:\n{flags['rule_id'].value_counts().sort_index()}")

if __name__ == "__main__":
    main()
